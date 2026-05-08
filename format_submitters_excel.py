# format_submitters_excel.py
# ==========================
# Step 2 of 2 (optional).
#
# Reads the submitters.csv produced by build_submitter_db.py and writes a
# styled Excel workbook: submitters_table.xlsx.
#
# Usage:
#     python format_submitters_excel.py --indir ./output
#
#     # Or point at specific files:
#     python format_submitters_excel.py --submitters ./output/submitters.csv --out ./output/submitters_table.xlsx
#
# Requirements:
#     pip install pandas openpyxl
 
import argparse
import os
import sys
 
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
 
 
# ── Colour palette ────────────────────────────────────────────────────────────
 
HEADER_FILL  = PatternFill("solid", start_color="1F4E79")   # dark navy
KEY_FILL     = PatternFill("solid", start_color="D6EAF8")   # light blue  (submitter_id column)
EXT_FILL     = PatternFill("solid", start_color="FEF9E7")   # light yellow (external/non-wright domains)
ALT_FILL     = PatternFill("solid", start_color="F5FBFF")   # very light blue (alternating rows)
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")
 
thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
 
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11, name="Arial")
KEY_FONT  = Font(name="Arial", size=10, bold=True)
DATA_FONT = Font(name="Arial", size=10)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_ALIGN = Alignment(vertical="top")
 
 
# ── Column layout ─────────────────────────────────────────────────────────────
 
# Each tuple: (csv_field_name, excel_header_label, column_width)
COLUMNS = [
    ("submitter_id", "submitter_id", 14),
    ("email",        "email",        38),
    ("display_name", "display_name", 28),
    ("username",     "username",     24),
    ("domain",       "domain",       28),
]
 
# Domain considered "internal" — rows with other domains get a highlight
INTERNAL_DOMAIN = "wright.edu"
 
 
# ── Cell helpers ──────────────────────────────────────────────────────────────
 
def header_cell(ws, value: str) -> WriteOnlyCell:
    # Build a fully styled header cell ready to append to a row.
    c = WriteOnlyCell(ws, value=value)
    c.font      = HDR_FONT
    c.fill      = HEADER_FILL
    c.alignment = HDR_ALIGN
    c.border    = BORDER
    return c
 
 
def data_cell(ws, value, fill=None, bold: bool = False) -> WriteOnlyCell:
    # Build a styled data cell. Converts any non-string value to string safely.
    v = "" if (value is None or (not isinstance(value, str) and pd.isna(value))) else str(value)
    c = WriteOnlyCell(ws, value=v)
    c.font      = Font(name="Arial", size=10, bold=bold)
    c.fill      = fill or WHITE_FILL
    c.alignment = TOP_ALIGN
    c.border    = BORDER
    return c
 
 
# ── Main ──────────────────────────────────────────────────────────────────────
 
def format_submitters(submitters_csv: str, output_xlsx: str) -> None:
    print(f"Reading {submitters_csv}...")
    df = pd.read_csv(submitters_csv, dtype=str).fillna("")
    print(f"  {len(df):,} submitters loaded")
 
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Submitters")
 
    # Set column widths
    for col_idx, (_, _, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
 
    # Freeze the header row so it stays visible when scrolling
    ws.freeze_panes = "A2"
 
    # Write the header row
    ws.append([header_cell(ws, label) for _, label, _ in COLUMNS])
 
    # Write the data rows
    total = 0
    for i, row in df.iterrows():
        # External submitters (non-wright.edu) get a yellow highlight
        is_external = row.get("domain", "") != INTERNAL_DOMAIN
        row_fill    = EXT_FILL if is_external else (ALT_FILL if i % 2 == 0 else WHITE_FILL)
 
        cells = []
        for field, _, _ in COLUMNS:
            val    = row.get(field, "")
            is_key = (field == "submitter_id")
            cells.append(
                data_cell(ws, val, fill=KEY_FILL if is_key else row_fill, bold=is_key)
            )
 
        ws.append(cells)
        total += 1
 
        if total % 200 == 0:
            print(f"  {total:,} rows written...")
 
    os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)
    wb.save(output_xlsx)
    print(f"\n✓ Saved: {output_xlsx}  ({total:,} submitters)")
    print("\nColumn guide:")
    print("  submitter_id  — primary key (SU000001 … SU000757)")
    print("  email         — full email address (lowercased)")
    print("  display_name  — 'First Last' derived from username (e.g. jane.smith -> Jane Smith)")
    print("  username      — part before the @")
    print("  domain        — part after the @")
    print(f"  Row colour    — white/blue = {INTERNAL_DOMAIN} accounts; yellow = external domains")
 
 
# ── CLI ───────────────────────────────────────────────────────────────────────
 
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert submitters.csv into a styled Excel workbook."
    )
    parser.add_argument(
        "--submitters", "-s",
        default=None,
        help="Path to submitters.csv (default: --indir/submitters.csv)"
    )
    parser.add_argument(
        "--indir",
        default="./output",
        help="Directory containing submitters.csv (default: ./output)"
    )
    parser.add_argument(
        "--out", "-o",
        default=None,
        help="Path for output .xlsx (default: --indir/submitters_table.xlsx)"
    )
    args = parser.parse_args()
 
    submitters_csv = args.submitters or os.path.join(args.indir, "submitters.csv")
    output_xlsx    = args.out        or os.path.join(args.indir, "submitters_table.xlsx")
 
    if not os.path.exists(submitters_csv):
        print(f"Error: submitters.csv not found at: {submitters_csv}", file=sys.stderr)
        print("Run build_submitter_db.py first.", file=sys.stderr)
        sys.exit(1)
 
    format_submitters(submitters_csv, output_xlsx)
