# format_authors_excel.py
# =======================
# Step 2 of 2 (optional).
#
# Reads the authors.csv produced by build_author_db.py and writes a styled
# Excel workbook: authors_table.xlsx.
#
# The records CSV (records.csv) is left as-is — at 89k+ rows it is too large
# for openpyxl to style efficiently, but opens cleanly in Excel or can be
# imported into any database.
#
# Usage:
#     python format_authors_excel.py --indir ./output --out ./output
#
#     # Or point at specific files:
#     python format_authors_excel.py --authors ./output/authors.csv --out ./output/authors_table.xlsx
#
# Requirements:
#     pip install pandas openpyxl
 
import argparse
import os
import sys
 
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
 
 
# ── Colour palette ────────────────────────────────────────────────────────────
 
HEADER_FILL = PatternFill("solid", start_color="1F4E79")   # dark navy
KEY_FILL    = PatternFill("solid", start_color="D6EAF8")   # light blue  (author_id column)
CORP_FILL   = PatternFill("solid", start_color="FEF9E7")   # light yellow (corporate authors)
ALT_FILL    = PatternFill("solid", start_color="F5FBFF")   # very light blue (alternating rows)
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")
 
thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
 
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11, name="Arial")
KEY_FONT  = Font(name="Arial", size=10, bold=True)
DATA_FONT = Font(name="Arial", size=10)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_ALIGN = Alignment(vertical="top")
 
 
# ── Helpers ───────────────────────────────────────────────────────────────────
 
def header_cell(ws, value: str) -> WriteOnlyCell:
    c = WriteOnlyCell(ws, value=value)
    c.font      = HDR_FONT
    c.fill      = HEADER_FILL
    c.alignment = HDR_ALIGN
    c.border    = BORDER
    return c
 
 
def data_cell(ws, value, fill=None, bold: bool = False) -> WriteOnlyCell:
    v = "" if (value is None or (not isinstance(value, str) and pd.isna(value))) else str(value)
    c = WriteOnlyCell(ws, value=v)
    c.font      = Font(name="Arial", size=10, bold=bold)
    c.fill      = fill or WHITE_FILL
    c.alignment = TOP_ALIGN
    c.border    = BORDER
    return c
 
 
# ── Column layout ─────────────────────────────────────────────────────────────
 
COLUMNS = [
    # (field_name,      display_header,   width)
    ("author_id",       "author_id",       12),
    ("display_name",    "display_name",    32),
    ("last_name",       "last_name",       20),
    ("first_name",      "first_name",      18),
    ("middle_name",     "middle_name",     14),
    ("suffix",          "suffix",          10),
    ("is_corporate",    "is_corporate",    12),
    ("corporate_name",  "corporate_name",  30),
    ("email",           "email",           32),
    ("institution",     "institution",     42),
]
 
 
# ── Main ──────────────────────────────────────────────────────────────────────
 
def format_authors(authors_csv: str, output_xlsx: str) -> None:
    print(f"Reading {authors_csv}...")
    df = pd.read_csv(authors_csv, dtype=str).fillna("")
    print(f"  {len(df):,} authors loaded")
 
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Authors")
 
    # Column widths
    for col_idx, (_, _, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
 
    # Freeze header row
    ws.freeze_panes = "A2"
 
    # Header row
    ws.append([header_cell(ws, display) for _, display, _ in COLUMNS])
 
    # Data rows
    total = 0
    for i, row in df.iterrows():
        is_corp  = row.get("is_corporate", "") == "Yes"
        row_fill = CORP_FILL if is_corp else (ALT_FILL if i % 2 == 0 else WHITE_FILL)
 
        cells = []
        for field, _, _ in COLUMNS:
            val  = row.get(field, "")
            is_key = (field == "author_id")
            cells.append(data_cell(ws, val, fill=KEY_FILL if is_key else row_fill, bold=is_key))
 
        ws.append(cells)
        total += 1
 
        if total % 5000 == 0:
            print(f"  {total:,} rows written...")
 
    os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)
    wb.save(output_xlsx)
    print(f"\n✓ Saved: {output_xlsx}  ({total:,} authors)")
    print("\nColumn guide:")
    print("  author_id       — primary key (AU000001 … AU028709)")
    print("  display_name    — 'Last, First Middle Suffix' or corporate name")
    print("  is_corporate    — Yes/No; corporate rows are highlighted yellow")
    print("  email/institution — enriched from best occurrence across all records")
 
 
# ── CLI ───────────────────────────────────────────────────────────────────────
 
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Convert authors.csv into a styled Excel workbook."
    )
    parser.add_argument(
        "--authors", "-a",
        default=None,
        help="Path to authors.csv (default: --indir/authors.csv)"
    )
    parser.add_argument(
        "--indir",
        default="./output",
        help="Directory containing authors.csv (default: ./output)"
    )
    parser.add_argument(
        "--out", "-o",
        default=None,
        help="Path for output .xlsx (default: --indir/authors_table.xlsx)"
    )
    args = parser.parse_args()
 
    authors_csv  = args.authors or os.path.join(args.indir, "authors.csv")
    output_xlsx  = args.out     or os.path.join(args.indir, "authors_table.xlsx")
 
    if not os.path.exists(authors_csv):
        print(f"Error: authors.csv not found at: {authors_csv}", file=sys.stderr)
        print("Run build_author_db.py first.", file=sys.stderr)
        sys.exit(1)
 
    format_authors(authors_csv, output_xlsx)
