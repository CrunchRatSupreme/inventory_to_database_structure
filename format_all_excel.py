# format_all_excel.py
# ===================
# Reads authors.csv and submitters.csv (produced by build_all_dbs.py) and
# writes a single styled Excel workbook with two sheets:
#
#   Sheet 1 — Authors    : one row per unique author
#   Sheet 2 — Submitters : one row per unique submitter
#
# Usage:
#     python format_all_excel.py --indir ./output
#
#     # Or point at specific files and output path:
#     python format_all_excel.py \
#         --authors    ./output/authors.csv \
#         --submitters ./output/submitters.csv \
#         --out        ./output/registry_tables.xlsx
#
# Requirements:
#     pip install pandas openpyxl

import argparse
import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Shared colour palette ─────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")  # dark navy   — header row
KEY_FILL    = PatternFill("solid", start_color="D6EAF8")  # light blue  — ID column
CORP_FILL   = PatternFill("solid", start_color="FEF9E7")  # light yellow — corporate authors
EXT_FILL    = PatternFill("solid", start_color="FEF9E7")  # light yellow — external submitters
ALT_FILL    = PatternFill("solid", start_color="F5FBFF")  # very light blue — alternating rows
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")

thin   = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HDR_FONT  = Font(bold=True, color="FFFFFF", size=11, name="Arial")
DATA_FONT = Font(name="Arial", size=10)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_ALIGN = Alignment(vertical="top")


# ── Shared column layouts ─────────────────────────────────────────────────────

# Authors sheet — each tuple: (csv_field_name, header_label, column_width)
AUTHOR_COLUMNS = [
    ("author_id",      "author_id",      12),
    ("display_name",   "display_name",   32),
    ("last_name",      "last_name",      20),
    ("first_name",     "first_name",     18),
    ("middle_name",    "middle_name",    14),
    ("suffix",         "suffix",         10),
    ("is_corporate",   "is_corporate",   12),
    ("corporate_name", "corporate_name", 30),
    ("email",          "email",          32),
    ("institution",    "institution",    42),
]

# Submitters sheet
SUBMITTER_COLUMNS = [
    ("submitter_id", "submitter_id", 14),
    ("email",        "email",        38),
    ("display_name", "display_name", 28),
    ("username",     "username",     24),
    ("domain",       "domain",       28),
]

# Submitters on this domain are considered internal; all others are highlighted
INTERNAL_DOMAIN = "wright.edu"


# ── Shared cell helpers ───────────────────────────────────────────────────────

def header_cell(ws, value: str) -> WriteOnlyCell:
    # Build a styled header cell — dark navy background, white bold text.
    c = WriteOnlyCell(ws, value=value)
    c.font      = HDR_FONT
    c.fill      = HEADER_FILL
    c.alignment = HDR_ALIGN
    c.border    = BORDER
    return c


def data_cell(ws, value, fill=None, bold: bool = False) -> WriteOnlyCell:
    # Build a styled data cell. Safely converts any value to a string.
    v = "" if (value is None or (not isinstance(value, str) and pd.isna(value))) else str(value)
    c = WriteOnlyCell(ws, value=v)
    c.font      = Font(name="Arial", size=10, bold=bold)
    c.fill      = fill or WHITE_FILL
    c.alignment = TOP_ALIGN
    c.border    = BORDER
    return c


def set_column_widths(ws, columns: list) -> None:
    # Set the width of each column from the layout definition.
    for col_idx, (_, _, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ── Sheet writers ─────────────────────────────────────────────────────────────

def write_authors_sheet(wb: Workbook, authors_csv: str) -> int:
    # Write the Authors sheet into the workbook.
    # Returns the number of rows written.
    print(f"  Reading {authors_csv}...")
    df = pd.read_csv(authors_csv, dtype=str).fillna("")
    print(f"  {len(df):,} authors loaded")

    ws = wb.create_sheet("Authors")
    set_column_widths(ws, AUTHOR_COLUMNS)
    ws.freeze_panes = "A2"

    # Header row
    ws.append([header_cell(ws, label) for _, label, _ in AUTHOR_COLUMNS])

    # Data rows
    total = 0
    for i, row in df.iterrows():
        # Corporate authors get a yellow highlight; others alternate white/light-blue
        is_corp  = row.get("is_corporate", "") == "Yes"
        row_fill = CORP_FILL if is_corp else (ALT_FILL if i % 2 == 0 else WHITE_FILL)

        cells = []
        for field, _, _ in AUTHOR_COLUMNS:
            is_key = (field == "author_id")
            cells.append(
                data_cell(ws, row.get(field, ""),
                          fill=KEY_FILL if is_key else row_fill,
                          bold=is_key)
            )
        ws.append(cells)
        total += 1

        if total % 5000 == 0:
            print(f"    {total:,} author rows written...")

    return total


def write_submitters_sheet(wb: Workbook, submitters_csv: str) -> int:
    # Write the Submitters sheet into the workbook.
    # Returns the number of rows written.
    print(f"  Reading {submitters_csv}...")
    df = pd.read_csv(submitters_csv, dtype=str).fillna("")
    print(f"  {len(df):,} submitters loaded")

    ws = wb.create_sheet("Submitters")
    set_column_widths(ws, SUBMITTER_COLUMNS)
    ws.freeze_panes = "A2"

    # Header row
    ws.append([header_cell(ws, label) for _, label, _ in SUBMITTER_COLUMNS])

    # Data rows
    total = 0
    for i, row in df.iterrows():
        # External (non-wright.edu) submitters get a yellow highlight
        is_external = row.get("domain", "") != INTERNAL_DOMAIN
        row_fill    = EXT_FILL if is_external else (ALT_FILL if i % 2 == 0 else WHITE_FILL)

        cells = []
        for field, _, _ in SUBMITTER_COLUMNS:
            is_key = (field == "submitter_id")
            cells.append(
                data_cell(ws, row.get(field, ""),
                          fill=KEY_FILL if is_key else row_fill,
                          bold=is_key)
            )
        ws.append(cells)
        total += 1

        if total % 200 == 0:
            print(f"    {total:,} submitter rows written...")

    return total


# ── Main ──────────────────────────────────────────────────────────────────────

def format_all(authors_csv: str, submitters_csv: str, output_xlsx: str) -> None:
    # Create a single workbook and write both sheets into it.
    wb = Workbook(write_only=True)

    print("\nWriting Authors sheet...")
    author_count = write_authors_sheet(wb, authors_csv)

    print("\nWriting Submitters sheet...")
    submitter_count = write_submitters_sheet(wb, submitters_csv)

    os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)
    wb.save(output_xlsx)

    print(f"\n✓ Saved: {output_xlsx}")
    print(f"  Sheet 'Authors'    : {author_count:,} rows")
    print(f"  Sheet 'Submitters' : {submitter_count:,} rows")
    print()
    print("Authors column guide:")
    print("  author_id       — primary key (AU000001 …)")
    print("  display_name    — 'Last, First Middle Suffix' or corporate name")
    print("  is_corporate    — Yes/No; corporate rows highlighted yellow")
    print("  email/institution — enriched from best occurrence across all records")
    print()
    print("Submitters column guide:")
    print("  submitter_id    — primary key (SU000001 …)")
    print("  email           — full email address (lowercased)")
    print("  display_name    — 'First Last' derived from username")
    print("  username        — part before the @")
    print("  domain          — part after the @")
    print(f"  Row colour      — white/blue = {INTERNAL_DOMAIN}; yellow = external domains")


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=(
            "Convert authors.csv and submitters.csv into a single styled "
            "Excel workbook with one sheet each."
        )
    )
    parser.add_argument(
        "--authors", "-a",
        default=None,
        help="Path to authors.csv (default: --indir/authors.csv)"
    )
    parser.add_argument(
        "--submitters", "-s",
        default=None,
        help="Path to submitters.csv (default: --indir/submitters.csv)"
    )
    parser.add_argument(
        "--indir",
        default="./output",
        help="Directory containing authors.csv and submitters.csv (default: ./output)"
    )
    parser.add_argument(
        "--out", "-o",
        default=None,
        help="Path for output .xlsx (default: --indir/registry_tables.xlsx)"
    )
    args = parser.parse_args()

    # Resolve file paths — use explicit args if given, otherwise fall back to indir
    authors_csv    = args.authors    or os.path.join(args.indir, "authors.csv")
    submitters_csv = args.submitters or os.path.join(args.indir, "submitters.csv")
    output_xlsx    = args.out        or os.path.join(args.indir, "registry_tables.xlsx")

    # Check both input files exist before doing any work
    missing = [f for f in (authors_csv, submitters_csv) if not os.path.exists(f)]
    if missing:
        for f in missing:
            print(f"Error: file not found: {f}", file=sys.stderr)
        print("Run build_all_dbs.py first.", file=sys.stderr)
        sys.exit(1)

    format_all(authors_csv, submitters_csv, output_xlsx)